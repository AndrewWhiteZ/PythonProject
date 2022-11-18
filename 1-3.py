import math
import os
import csv
import itertools
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import openpyxl as excel
import pdfkit

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

dic_naming = {
    'name': 'Название',
    'salary': 'Оклад',
    'salary_from': 'Нижняя граница вилки оклада',
    'salary_to': 'Верхняя граница вилки оклада',
    'salary_currency': 'Идентификатор валюты оклада',
    'area_name': 'Название региона',
    'published_at': 'Дата публикации вакансии'
}


def format_columns_width(worksheet):
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(str(cell.value)))

    for column, value in dims.items():
        worksheet.column_dimensions[column].width = value + 2


def set_header(worksheet, columns_title, bd):
    for index, title in enumerate(columns_title):
        cell = worksheet.cell(column=index + 1, row=1, value=title)
        if title:
            cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            cell.font = Font(bold=True)


def write_data(worksheet, data, bd, first_col_number=1, first_row_number=2, percentage_col_number=-1):
    for i, row in enumerate(data):
        for j, value in enumerate(row):
            cell = worksheet.cell(column=j + first_col_number, row=i + first_row_number, value=data[i][j])
            cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            if j + 1 == percentage_col_number:
                cell.number_format = FORMAT_PERCENTAGE_00


class Report(object):
    def __init__(self, ws_titles_list, border_style):
        self.ws_titles_headers = ws_titles_list
        self.border_style = border_style

    def generate_excel(self, dic_list):
        wb = excel.Workbook()
        ws_list = []
        bd = self.border_style

        by_year_data = []
        by_city_data_salary = []
        by_city_data_ratio = []

        for year in list(dic_list[0].keys()):
            by_year_data.append([year, dic_list[0][year], dic_list[2][year],
                                 dic_list[1][year], dic_list[3][year]])

        for city in list(dic_list[4].keys()):
            by_city_data_salary.append([city, dic_list[4][city]])

        for city in list(dic_list[5].keys()):
            by_city_data_ratio.append([city, dic_list[5][city]])

        for index, title in enumerate(list(self.ws_titles_headers.keys())):
            if index == 0:
                worksheet = wb.active
                worksheet.title = title
            else:
                worksheet = wb.create_sheet(title)
            ws_list.append(worksheet)
            set_header(worksheet, self.ws_titles_headers[title], bd)

        write_data(ws_list[0], by_year_data, bd)
        write_data(ws_list[1], by_city_data_salary, bd)
        write_data(ws_list[1], by_city_data_ratio, bd, 4, 2, 2)

        format_columns_width(ws_list[0])
        format_columns_width(ws_list[1])
        wb.save('report.xlsx')

    def generate_image(self, dic_list):
        plt.rc('axes', labelsize=8)
        plt.rc('xtick', labelsize=8)
        plt.rc('ytick', labelsize=8)
        plt.rc('legend', fontsize=8)

        labels = list(dic_list[0].keys())
        cities = list(dic_list[4].keys())

        one_cities = ['\n'.join(x.split(' ')) for x in cities]
        new_cities = ['-\n'.join(x.split('-')) for x in one_cities]

        x = np.arange(len(labels))
        y = np.arange(len(cities))
        width = 0.35

        fig, axs = plt.subplots(2, 2, figsize=(19.2, 14.4))
        axs[0, 0].bar(x - width / 2, list(dic_list[0].values()), width, label='средняя з/п')
        axs[0, 0].bar(x + width / 2, list(dic_list[2].values()), width, label=f'з/п {profession}')
        axs[0, 0].set_title('Уровень зарплат по годам')
        axs[0, 0].set_xticks(x, labels, rotation=90)
        axs[0, 0].legend()
        axs[0, 0].grid(axis='y')

        axs[0, 1].bar(x - width / 2, list(dic_list[1].values()), width, label='Количество вакансий')
        axs[0, 1].bar(x + width / 2, list(dic_list[3].values()), width,
                      label=f'Количество вакансий {profession}')
        axs[0, 1].set_title('Количество вакансий по годам')
        axs[0, 1].set_xticks(x, labels, rotation=90)
        axs[0, 1].legend()
        axs[0, 1].grid(axis='y')

        axs[1, 0].barh(cities, list(dic_list[4].values()))
        axs[1, 0].set_title('Уровень зарплат по городам')
        axs[1, 0].set_yticks(y, new_cities)
        axs[1, 0].invert_yaxis()
        axs[1, 0].grid(axis='x')

        others_ratio = 1 - sum(list(dic_list[5].values()))

        axs[1, 1].pie([others_ratio] + (list(dic_list[5].values())), labels=['Другие'] + cities)
        axs[1, 1].set_title('Доля вакансий по городам')
        axs[1, 1].axis('equal')

        fig.tight_layout()
        plt.savefig(r"C:\Users\Andrew\PycharmProjects\YandexContest2\graph.png", dpi=200)

    def generate_pdf(self):
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("html_template.html")
        book = load_workbook("report.xlsx")
        ws_list = book.worksheets
        pdf_template = template.render({'profession': profession}, first_ws=ws_list[0], second_ws=ws_list[1])
        config = pdfkit.configuration(wkhtmltopdf=r'E:\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": True})


class Salary(object):
    def __init__(self, salary_f, salary_t, salary_c, salary_g='Нет'):
        self.salary_from = salary_f
        self.salary_to = salary_t
        self.salary_gross = salary_g
        self.salary_currency = salary_c

    def convert(self):
        return (int(float(self.salary_from)) + int(float(self.salary_to))) / 2 * currency_to_rub[self.salary_currency]


class Vacancy(object):
    def __init__(self, vacancy_row):
        if len(vacancy_row) == 6:
            self.name = vacancy_row[0]
            self.salary = Salary(vacancy_row[1], vacancy_row[2], vacancy_row[3])
            self.area_name = vacancy_row[4]
            self.published_at = vacancy_row[5]
        else:
            self.name = vacancy_row[0]
            self.description = vacancy_row[1]
            self.key_skills = vacancy_row[2]
            self.experience_id = vacancy_row[3]
            self.premium = vacancy_row[4]
            self.employer_name = vacancy_row[5]
            self.salary = Salary(vacancy_row[6], vacancy_row[7], vacancy_row[9], vacancy_row[8])
            self.area_name = vacancy_row[10]
            self.published_at = vacancy_row[11]


class DataSet(object):
    def __init__(self, file_path):
        vacancies_objs = []
        with open(file_path, encoding="utf-8-sig") as File:
            if os.stat(file).st_size == 0:
                self.vacancies = 'Пустой файл'
            else:
                reader = csv.reader(File, delimiter=',', quotechar='"')
                for vacancy in reader:
                    if all(vacancy):
                        vacancies.append(vacancy)
                        if len(vacancy) < len(vacancies[0]):
                            vacancies.remove(vacancy)

            if len(vacancies) == 1:
                self.vacancies = 'Нет данных'
            elif len(vacancies) != 0:
                for vacancy in vacancies[1::]:
                    vacancies_objs.append(Vacancy(vacancy))
                self.vacancies = vacancies_objs

    # Получаем датасет распределения зарплаты по годам
    def salary_by_year(self):
        for vacancy in self.vacancies:
            if int(vacancy.published_at[:4]) not in list(dic_salaries.keys()):
                dic_salaries.update({int(vacancy.published_at[:4]): 0})
                dic_counter.update({int(vacancy.published_at[:4]): 0})
            dic_salaries[int(vacancy.published_at[:4])] += vacancy.salary.convert()
            dic_counter[int(vacancy.published_at[:4])] += 1
        self.years = (dic_salaries.keys())
        for year in self.years:
            dic_salaries.update({year: math.floor(dic_salaries[year] / dic_counter[year])})
            dic_counter.update({year: math.floor(dic_counter[year])})
        return [dic_salaries, dic_counter]

    # Получаем датасет распределения зарплаты по профессиям
    def salary_by_profession(self):
        for vacancy in self.vacancies:
            if profession in vacancy.name:
                if int(vacancy.published_at[:4]) not in list(dic_salaries_by_profession.keys()):
                    dic_salaries_by_profession.update({int(vacancy.published_at[:4]): 0})
                    dic_counter_by_profession.update({int(vacancy.published_at[:4]): 0})
                dic_salaries_by_profession[int(vacancy.published_at[:4])] += vacancy.salary.convert()
                dic_counter_by_profession[int(vacancy.published_at[:4])] += 1
        if len(dic_salaries_by_profession) == 0:
            dic_salaries_by_profession.clear()
            dic_counter_by_profession.clear()
            dic_salaries_by_profession.update(dict({x: 0 for x in self.years}))
            dic_counter_by_profession.update(dict({x: 0 for x in self.years}))
            return [dic_salaries_by_profession, dic_counter_by_profession]
        for year in self.years:
            dic_salaries_by_profession.update({year: math.floor(dic_salaries_by_profession[year] /
                                                                dic_counter_by_profession[year])})
            dic_counter_by_profession.update({year: dic_counter_by_profession[year]})
        return [dic_salaries_by_profession, dic_counter_by_profession]

    # Получаем датасет распределения зарплаты по городам
    def salary_by_area(self):
        for vacancy in self.vacancies:
            if vacancy.area_name not in list(dic_cities_salary.keys()):
                dic_cities_salary.update({vacancy.area_name: 0})
                dic_cities_ratio.update({vacancy.area_name: 0})
            dic_cities_salary[vacancy.area_name] += vacancy.salary.convert()
            dic_cities_ratio[vacancy.area_name] += 1
        length = len(self.vacancies)
        for city in list(dic_cities_salary.keys()):
            if dic_cities_ratio[city] >= math.floor(length * 0.01):
                cities_list.append(city)
                new_dic_cities_salary.update({city: dic_cities_salary[city]})
                new_dic_cities_ratio.update({city: dic_cities_ratio[city]})
            else:
                del dic_cities_salary[city]
                del dic_cities_ratio[city]
        for city in cities_list:
            dic_cities_salary.update({city: math.floor(new_dic_cities_salary[city] / new_dic_cities_ratio[city])})
            dic_cities_ratio.update({city: round(new_dic_cities_ratio[city] / length, 4)})
        return [dict(itertools.islice(sorted(dic_cities_salary.items(), key=lambda item: item[1], reverse=True), 10)),
                dict(itertools.islice(sorted(dic_cities_ratio.items(), key=lambda item: item[1], reverse=True), 10))]

    def get_full_dict_list(self):
        return data.salary_by_year() + data.salary_by_profession() + data.salary_by_area()


vacancies = []
cities_list = []
dic_salaries = {}
dic_counter = {}
dic_salaries_by_profession = {}
dic_counter_by_profession = {}
dic_cities_salary = {}
dic_cities_ratio = {}
new_dic_cities_salary = {}
new_dic_cities_ratio = {}


file = input('Введите название файла: ')
profession = input('Введите название профессии: ')
data = DataSet(file)
full_dict_list = data.get_full_dict_list()

print(f'Динамика уровня зарплат по годам: {full_dict_list[0]}')
print(f'Динамика количества вакансий по годам: {full_dict_list[1]}')
print(f'Динамика уровня зарплат по годам для выбранной профессии: {full_dict_list[2]}')
print(f'Динамика количества вакансий по годам для выбранной профессии: {full_dict_list[3]}')
print(f'Уровень зарплат по городам (в порядке убывания): {full_dict_list[4]}')
print(f'Доля вакансий по городам (в порядке убывания): {full_dict_list[5]}')

report = Report({'Статистика по годам': ['Год', 'Средняя зарплата', f'Средняя зарплата - {profession}',
                                            'Количество вакансий', f'Количество вакансий - {profession}'],
                 'Статистика по городам': ['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']},
                Side(style='thin', color="000000"))

report.generate_excel(full_dict_list)
report.generate_image(full_dict_list)
report.generate_pdf()
